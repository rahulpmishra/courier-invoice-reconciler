#!/usr/bin/env python3
"""Generate the synthetic invoice + register used by the test suites.

The real files this project was built against contain customer phone numbers,
receiver names and bank details, so they are gitignored. These stand-ins carry
no personal data but reproduce every structural quirk the parsers must handle:

  * two invoices in one PDF, each with its own period
  * an "Additional Charges" table with a Consignee column that the "Standard"
    tables do not have, so the column map has to be read per page
  * a continuation page with rows but no header row
  * a section banner printed at the foot of the *previous* page
  * back-billed rows dated before the invoice period
  * a register mixing `4-June-26`, `4-Jun-2026`, real Excel dates and one
    unparseable value
  * a 13-digit typo of a real AWB, a duplicate booking, and AWBs from another
    carrier that appear on no invoice

Run:  python tools/make_fixtures.py
"""

import csv
import datetime
import os
import random
import zlib

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "tests", "fixtures")

CUSTOMER = "D00000 - ACME LOGISTICS (SAMPLE BRANCH)"
DESTS = ["DELHI", "MUMBAI", "PUNE", "KOLKATA", "CHENNAI", "JAIPUR", "INDORE",
         "NAGPUR", "SURAT", "LUCKNOW", "BHOPAL", "PATNA", "KOCHI", "RAIPUR"]

# Column x positions, lifted from the real invoice so the geometry matches.
STANDARD_COLS = [("SNo.", 39), ("Date", 74.64), ("AwbNo", 136.01), ("Origin", 212.61),
                 ("Dest", 257.56), ("Mode", 399.49), ("PCS", 438.49),
                 ("Weight", 471.38), ("Amount", 517.1)]
ADDL_COLS = [("SNo.", 38.25), ("Date", 68.1), ("AwbNo", 134.77), ("Origin", 220.03),
             ("Dest", 266.67), ("Mode", 364.42), ("Consignee", 397.54),
             ("PCS", 450.95), ("Weight", 478.12), ("Amount", 517.8)]

ROWS_PER_PAGE = 40
FIRST_ROW_Y = 702.64
ROW_PITCH = 12.77
BS = chr(92)


def build_rows(rng):
    """Ground truth: every consignment line the PDF will contain."""
    rows = []
    seq = [900461102680]   # varied digits, like real AWBs: a long run of
                           # zeros makes several neighbours one-deletion apart

    def make(date, invoice, section):
        seq[0] += rng.randint(1, 4)
        return {
            "invoice_no": invoice, "section": section, "date": date.strftime("%d/%m/%Y"),
            "awb": str(seq[0]), "origin": "DELHI", "dest": rng.choice(DESTS),
            "mode": rng.choice("SA"), "pcs": 1 if rng.random() < 0.9 else 2,
            "weight": round(rng.uniform(0.05, 9.5), 3),
            "amount": round(rng.uniform(18, 950), 2),
        }

    for day in (12, 18, 18, 25, 29):
        rows.append(make(datetime.date(2026, 5, day), "EXM/1000/A0001", "Additional Charges"))
    for i in range(60):
        rows.append(make(datetime.date(2026, 6, 1 + (i * 15) // 60), "EXM/1000/A0001", "Standard"))
    for i in range(55):
        rows.append(make(datetime.date(2026, 6, 16 + (i * 15) // 55), "EXM/1000/A0002", "Standard"))
    return rows


def build_register(rows, rng):
    """The shop's booking register: a partial, messier view of the same month."""
    june = [r for r in rows if r["section"] == "Standard"]
    matched = june[:40]
    out = []

    def entry(date, awb, dest):
        return {"DATE": date, "AWB NO.": awb,
                "SENDER NAME": "SAMPLE TRADER " + str(rng.randint(1, 20)),
                "ADDRESS": "SAMPLE ADDRESS",
                "RECEIVER NAME": "SAMPLE RECEIVER " + str(rng.randint(1, 99)),
                "PIN CODE": rng.randint(110001, 799999), "DESTINATION": dest,
                "WEIGHT IN KG": round(rng.uniform(0.1, 9.0), 2), "MODE": "SURFACE"}

    for i, r in enumerate(matched):
        day, month, year = r["date"].split("/")
        date = (str(int(day)) + "-June-26" if i % 2
                else datetime.datetime(int(year), int(month), int(day)))
        out.append(entry(date, r["awb"], r["dest"]))

    out.append(entry("7-June-26", matched[0]["awb"], matched[0]["dest"]))   # duplicate booking
    # A mistyped AWB. It is taken from a consignment the shop never booked
    # correctly, so the invoice line stays unmatched and each side can point at
    # the other - which is how the real data behaved.
    typo = june[45]["awb"]
    out.append(entry("9-June-26", typo[:4] + typo[3] + typo[4:], "PUNE"))
    for i, awb in enumerate(("770000000001", "770000000002", "770000000003")):
        out.append(entry(str(11 + i) + "-Jun-2026", awb, "SURAT"))          # another carrier
    out.append(entry("n/a", "770000000004", "INDORE"))                      # unreadable date

    for i in range(25):                                                     # next month, ignored
        out.append(entry(str(1 + i % 25) + "-July-2026", str(880000000000 + i), rng.choice(DESTS)))
    return out


def esc(text):
    text = str(text).replace(BS, BS + BS)
    return text.replace("(", BS + "(").replace(")", BS + ")")


def cell(x, y, text, size=7.19):
    return "q BT /F1 " + str(size) + " Tf " + str(x) + " " + str(y) + \
           " Td 0 0 0 rg (" + esc(text) + ")Tj ET Q\n"


def summary_page(inv, frm, to, invoiced, trailing_banner=None):
    parts = [cell(39, 800, "TAX INVOICE", 11), cell(39, 780, "Billed By", 8),
             cell(39, 768, "Example Couriers Pvt. Ltd."),
             cell(39, 756, "SAMPLE ADDRESS, NEW DELHI"),
             cell(39, 730, "Billed To"), cell(39, 718, "ACME LOGISTICS (SAMPLE BRANCH)"),
             cell(39, 706, "Code : D00000")]
    labels = (("Invoice Date :", invoiced), ("From Date :", frm),
              ("Date To :", to), ("Invoice No.:", inv))
    for i, (label, value) in enumerate(labels):
        y = 768 - i * 14
        parts.append(cell(260, y, label))
        parts.append(cell(513, y, value))
    if trailing_banner:
        # Section banner printed at the foot of the previous page, exactly as
        # the real invoice does - the parser has to carry it forward.
        parts.append(cell(39, 60, trailing_banner))
        parts.append(cell(477, 60, "Page No : 1"))
    return "".join(parts)


KEYS = {"SNo.": "sno", "Date": "date", "AwbNo": "awb", "Origin": "origin", "Dest": "dest",
        "Mode": "mode", "PCS": "pcs", "Weight": "weight", "Amount": "amount"}


def detail_page(section, page_no, invoice, cols, rows, with_header=True):
    parts = []
    if with_header:
        parts.append(cell(39, 745.8, section))
        parts.append(cell(477, 745.8, "Page No : " + str(page_no)))
        parts.append(cell(39, 731.53, CUSTOMER + " | " + invoice + " | Fuel(%).18.00 | Dev(%). 5.00"))
        for name, x in cols:
            parts.append(cell(x, 717, name, 7.5))
    for i, row in enumerate(rows):
        y = FIRST_ROW_Y - i * ROW_PITCH
        for name, x in cols:
            if name == "Consignee":
                continue                       # present as a column, always blank
            if name == "Weight":
                value = "%.3f" % row["weight"]
            elif name == "Amount":
                value = "%.2f" % row["amount"]
            else:
                value = row[KEYS[name]]
            parts.append(cell(x, y, value))
    return "".join(parts)


def write_pdf(path, pages):
    objects = ["<</Type/Catalog/Pages 2 0 R>>", None,
               "<</Type/Font/Subtype/Type1/BaseFont/Helvetica/Encoding/WinAnsiEncoding>>"]
    kids = []
    for content in pages:
        stream = zlib.compress(content.encode("latin-1"))
        objects.append(("stream", "<</Length " + str(len(stream)) + "/Filter/FlateDecode>>", stream))
        contents_id = len(objects)
        objects.append("<</Type/Page/Parent 2 0 R/MediaBox[0 0 595 842]"
                       "/Resources<</Font<</F1 3 0 R>>>>"
                       "/Contents " + str(contents_id) + " 0 R>>")
        kids.append(len(objects))
    objects[1] = ("<</Type/Pages/Kids[" + " ".join(str(k) + " 0 R" for k in kids) +
                  "]/Count " + str(len(kids)) + ">>")

    body, offsets = bytearray(b"%PDF-1.4\n"), []
    for i, obj in enumerate(objects, start=1):
        offsets.append(len(body))
        body += (str(i) + " 0 obj\n").encode()
        if isinstance(obj, tuple):
            body += obj[1].encode() + b"\nstream\n" + obj[2] + b"\nendstream"
        else:
            body += obj.encode()
        body += b"\nendobj\n"

    xref = len(body)
    body += ("xref\n0 " + str(len(objects) + 1) + "\n0000000000 65535 f \n").encode()
    for off in offsets:
        body += ("%010d 00000 n \n" % off).encode()
    body += ("trailer\n<</Size " + str(len(objects) + 1) + "/Root 1 0 R>>\n" +
             "startxref\n" + str(xref) + "\n%%EOF\n").encode()
    with open(path, "wb") as fh:
        fh.write(body)


def main():
    rng = random.Random(20260825)
    os.makedirs(OUT, exist_ok=True)
    rows = build_rows(rng)
    addl = [r for r in rows if r["section"] == "Additional Charges"]
    inv1 = [r for r in rows if r["invoice_no"] == "EXM/1000/A0001" and r["section"] == "Standard"]
    inv2 = [r for r in rows if r["invoice_no"] == "EXM/1000/A0002"]
    for chunk in (addl, inv1, inv2):
        for n, row in enumerate(chunk, start=1):
            row["sno"] = n

    pages = [
        summary_page("EXM/1000/A0001", "01/06/2026", "15/06/2026", "16/06/2026"),
        detail_page("Additional Charges", 1, "EXM/1000/A0001", ADDL_COLS, addl),
        detail_page("Standard", 1, "EXM/1000/A0001", STANDARD_COLS, inv1[:ROWS_PER_PAGE]),
        # Continuation page: rows but no header row, so the column map must carry.
        detail_page("Standard", 2, "EXM/1000/A0001", STANDARD_COLS, inv1[ROWS_PER_PAGE:],
                    with_header=False),
        summary_page("EXM/1000/A0002", "16/06/2026", "30/06/2026", "01/07/2026",
                     trailing_banner="Standard"),
        detail_page("Standard", 1, "EXM/1000/A0002", STANDARD_COLS, inv2[:ROWS_PER_PAGE]),
        detail_page("Standard", 2, "EXM/1000/A0002", STANDARD_COLS, inv2[ROWS_PER_PAGE:]),
    ]
    write_pdf(os.path.join(OUT, "sample-invoice.pdf"), pages)

    page_of = {}
    for n, chunk in ((2, addl), (3, inv1[:ROWS_PER_PAGE]), (4, inv1[ROWS_PER_PAGE:]),
                     (6, inv2[:ROWS_PER_PAGE]), (7, inv2[ROWS_PER_PAGE:])):
        for row in chunk:
            page_of[row["awb"]] = n

    ordered = addl + inv1 + inv2
    with open(os.path.join(OUT, "expected-rows.csv"), "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["invoice_no", "section", "page", "sno", "date", "awb",
                    "origin", "dest", "mode", "pcs", "weight", "amount"])
        for row in ordered:
            w.writerow([row["invoice_no"], row["section"], page_of[row["awb"]], row["sno"],
                        row["date"], row["awb"], row["origin"], row["dest"], row["mode"],
                        row["pcs"], "%.3f" % row["weight"], "%.2f" % row["amount"]])

    register = build_register(rows, rng)
    fields = list(register[0].keys())
    with open(os.path.join(OUT, "sample-register.csv"), "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fields)
        w.writeheader()
        for r in register:
            row = dict(r)
            if isinstance(row["DATE"], datetime.datetime):
                row["DATE"] = str(row["DATE"].day) + row["DATE"].strftime("-%B-%y")
            w.writerow(row)

    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Register"
        ws.append(fields)
        for r in register:
            ws.append([r[f] for f in fields])
        wb.save(os.path.join(OUT, "sample-register.xlsx"))

        # The same rows split across tabs, the way a register drifts once
        # someone starts a new sheet each month. Reconciling this must give
        # exactly the same answer as the flat file above. The third sheet has
        # no AWB column and must be ignored rather than parsed as bookings.
        split = len(register) - 25
        wb = Workbook()
        for name, chunk in (("June", register[:split]), ("July", register[split:])):
            ws = wb.create_sheet(name)
            ws.append(fields)
            for r in chunk:
                ws.append([r[f] for f in fields])
        rates = wb.create_sheet("Rates")
        rates.append(["ZONE", "PER KG"])
        for zone, rate in (("NORTH", 42), ("SOUTH", 55), ("EAST", 48)):
            rates.append([zone, rate])
        del wb["Sheet"]
        wb.save(os.path.join(OUT, "sample-register-multisheet.xlsx"))
        xlsx = "sample-register.xlsx, sample-register-multisheet.xlsx"
    except ImportError:
        xlsx = "(skipped - openpyxl not installed)"

    print("invoice   %4d rows across %d pages -> sample-invoice.pdf" % (len(ordered), len(pages)))
    print("register  %4d rows -> sample-register.csv, %s" % (len(register), xlsx))


if __name__ == "__main__":
    main()
